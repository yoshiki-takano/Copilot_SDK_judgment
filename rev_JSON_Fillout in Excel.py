# write_json_to_excel.py
import argparse
import json
import os
import re
import sys
from typing import Any, Dict, List, Tuple, Union

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

Scalar = Union[str, int, float, bool, None]
JSONType = Union[Dict[str, Any], List[Any], Scalar]

def add_stage_prefix(record: Dict[str, Any], prefix: str) -> Dict[str, Any]:
    return {f"{prefix}_{k}": v for k, v in record.items()}

def strip_code_fences(s: str) -> str:
    if not isinstance(s, str):
        return s
    fence_re = re.compile(r"^\s*```(?:json|JSON)?\s*([\s\S]*?)\s*```\s*$")
    m = fence_re.match(s.strip())
    return m.group(1) if m else s


def parse_json_input(inp: str) -> JSONType:
    if inp == "-":
        raw = sys.stdin.read()
    elif os.path.isfile(inp):
        with open(inp, "r", encoding="utf-8") as f:
            raw = f.read()
    else:
        raw = inp
    raw = strip_code_fences(raw).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        return {"_raw_text": raw, "_json_error": str(e)}


def is_list_of_dicts(x: Any) -> bool:
    return isinstance(x, list) and all(isinstance(i, dict) for i in x)


def is_list_of_scalars(x: Any) -> bool:
    return isinstance(x, list) and all(not isinstance(i, (dict, list)) for i in x)


def flatten_dict(d: Dict[str, Any], parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
    items: List[Tuple[str, Any]] = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def ensure_sheet(wb: Workbook, sheetname: str):
    if sheetname in wb.sheetnames:
        return wb[sheetname]
    else:
        ws = wb.create_sheet(title=sheetname)
        return ws


def read_header(ws, header_row: int, start_col: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    col = start_col
    while True:
        val = ws.cell(row=header_row, column=col).value
        if val is None or str(val).strip() == "":
            break
        header_map[str(val)] = col
        col += 1
    return header_map


def write_header(ws, header_row: int, start_col: int, header_map: Dict[str, int], new_fields: List[str]) -> Dict[str, int]:
    if not header_map:
        col = start_col
    else:
        col = max(header_map.values()) + 1
    for field in new_fields:
        if field not in header_map:
            ws.cell(row=header_row, column=col, value=field)
            header_map[field] = col
            col += 1
    return header_map


def to_excel_scalar(val: Any, delimiter: str) -> Scalar:
    if isinstance(val, list):
        if is_list_of_scalars(val):
            return delimiter.join("" if v is None else str(v) for v in val)
        else:
            return json.dumps(val, ensure_ascii=False)
    elif isinstance(val, dict):
        return json.dumps(val, ensure_ascii=False)
    else:
        return val


def expand_list_cols(field: str, values: List[Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for i, v in enumerate(values, start=1):
        out[f"{field}[{i}]"] = v
    return out


def union_field_names(records: List[Dict[str, Any]]) -> List[str]:
    fields = []
    seen = set()
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                fields.append(k)
    return fields


def write_row(ws, row_idx: int, header_map: Dict[str, int], record: Dict[str, Any], delimiter: str):
    for k, v in record.items():
        col_idx = header_map[k]
        ws.cell(row=row_idx, column=col_idx, value=to_excel_scalar(v, delimiter))


def auto_shape_record_from_dict(
    data: Dict[str, Any],
    list_style: str,
    delimiter: str
) -> Dict[str, Any]:
    flat = flatten_dict(data)
    out: Dict[str, Any] = {}
    pending_spill: Dict[str, Any] = {}
    for k, v in flat.items():
        if isinstance(v, list) and list_style == "cols" and is_list_of_scalars(v):
            pending_spill.update(expand_list_cols(k, v))
        else:
            out[k] = v
    out.update(pending_spill)
    return out


def find_first_empty_column(ws) -> int:
    max_col = ws.max_column if ws.max_column is not None else 1
    max_row = ws.max_row if ws.max_row is not None else 1
    for c in range(1, max_col + 1):
        has_any = False
        for r in range(1, max_row + 1):
            if ws.cell(row=r, column=c).value is not None:
                has_any = True
                break
        if not has_any:
            return c
    return max_col + 1


def find_existing_header_start(ws, header_row: int) -> Union[int, None]:
    """
    ヘッダー行(header_row)の左端の非空セルの列番号を返す。無ければNone。
    """
    max_col = ws.max_column if ws.max_column is not None else 1
    for c in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None and str(val).strip() != "":
            return c
    return None


def main():
    parser = argparse.ArgumentParser(
        description="可変JSON→Excel書き込みツール（ヘッダー自動・構造自動判定対応）",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument("json_input", nargs="?", help="JSON文字列 / .jsonファイルパス / '-'(stdin)")
    parser.add_argument("xlsx_path", nargs="?", help="出力/追記するExcelファイルパス")
    parser.add_argument("sheetname", nargs="?", help="シート名")
    parser.add_argument("--start-row", type=int, default=2, help="データ書き込み開始行（ヘッダーの次行を想定）")
    parser.add_argument("--start-col", type=int, default=8, help="データ書き込み開始列（1始まり：初回の下限）")
    parser.add_argument("--header-row", type=int, default=1, help="ヘッダー行番号（1始まり）")
    parser.add_argument("--orient", choices=["auto", "row", "rows"], default="auto",
                        help="トップレベル構造の扱い：auto=自動判定, row=dictを1行, rows=list[dict]を複数行")
    parser.add_argument("--list-style", choices=["join", "cols"], default="join",
                        help="list[scalar] の扱い：join=1セル結合, cols=列スピル")
    parser.add_argument("--delimiter", default="; ", help="join時の区切り文字")
    parser.add_argument("--create-if-missing", action="store_true", help="Excelが無い場合に新規作成")
    parser.add_argument("--quiet", action="store_true", help="ログ最小化")

    use_legacy = (len(sys.argv) == 6 and sys.argv[1] not in ("--help", "-h"))

    if use_legacy:
        json_string = sys.argv[1]
        xlsx_path = sys.argv[2]
        start_row = int(sys.argv[3])
        start_col = int(sys.argv[4])
        sheetname = sys.argv[5]
        header_row = 1
        orient = "auto"
        list_style = "join"
        delimiter = "; "
        create_if_missing = False
        quiet = False
    else:
        args = parser.parse_args()
        if not args.json_input or not args.xlsx_path or not args.sheetname:
            parser.print_help()
            sys.exit(1)
        json_string = args.json_input
        xlsx_path = args.xlsx_path
        sheetname = args.sheetname
        start_row = args.start_row
        start_col = args.start_col
        header_row = args.header_row
        orient = args.orient
        list_style = args.list_style
        delimiter = args.delimiter
        create_if_missing = args.create_if_missing
        quiet = args.quiet

    if not quiet:
        print("JSON処理開始")

    data = parse_json_input(json_string)

    # Excel の用意
    try:
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
        else:
            if create_if_missing:
                wb = Workbook()
                default = wb.active
                wb.remove(default)
            else:
                print(f"エラー: ファイル '{xlsx_path}' が見つかりません（--create-if-missing で新規作成可）。")
                sys.exit(1)
        ws = ensure_sheet(wb, sheetname)
    except Exception as e:
        print(f"Excelファイル '{xlsx_path}' の読み込み/作成中にエラーが発生しました: {e}")
        sys.exit(1)

    # === 列起点の固定化 ===
    existing_header_start = find_existing_header_start(ws, header_row)
    if existing_header_start is not None:
        auto_empty_col = existing_header_start
        if not quiet:
            print(f"既存ヘッダー検出: 列 {start_col} を起点として固定します")
    else:
        auto_empty_col = find_first_empty_column(ws)
        start_col =  start_col #max(start_col, auto_empty_col)
        if not quiet:
            print(f"ヘッダー未検出: 最初の空列 {auto_empty_col} と指定 {start_col} の大きい方 → 開始列 {start_col}")

    # 既存ヘッダー読み取り
    header_map = read_header(ws, header_row=header_row, start_col=start_col)

    # 書き込みレコード準備
    records: List[Dict[str, Any]] = []

    if isinstance(data, dict):
        record = auto_shape_record_from_dict(data, list_style=list_style, delimiter=delimiter)
        
        # ★ここ追加
        # stage2想定（必要なら引数化してもOK）
        record = add_stage_prefix(record, "2")

        records = [record]
        if not quiet:
            print("トップレベル: dict → 1行出力")
    elif is_list_of_dicts(data):
        if orient == "row":
            records = [{"_raw": data}]
            if not quiet:
                print("トップレベル: list[dict] だが --orient=row → _raw にJSON格納")
        else:
            if not quiet:
                print(f"トップレベル: list[dict] → {len(data)} 行に出力")
            for item in data:
                rec = auto_shape_record_from_dict(item, list_style=list_style, delimiter=delimiter)
                records.append(rec)
    elif is_list_of_scalars(data):
        if list_style == "cols":
            rec = expand_list_cols("value", data)
        else:
            rec = {"value": delimiter.join("" if v is None else str(v) for v in data)}
        records = [rec]
        if not quiet:
            print("トップレベル: list[scalar]")
    else:
        records = [{"_raw": data}]
        if not quiet:
            print("トップレベル: その他（_raw にJSON格納）")

    # ヘッダー更新
    all_fields = union_field_names(records)
    header_map = write_header(ws, header_row=header_row, start_col=start_col, header_map=header_map, new_fields=all_fields)

    # 書き込み開始行（固定で start_row から）
    for i, rec in enumerate(records):
        row_idx = start_row + i
        write_row(ws, row_idx, header_map, rec, delimiter=delimiter)
        if not quiet:
            print(f"行 {row_idx} にレコードを書き込み")

    # 保存
    try:
        wb.save(xlsx_path)
        if not quiet:
            print(f"ファイル '{xlsx_path}' を保存しました。")
    except Exception as e:
        print(f"ファイルの保存中にエラーが発生しました: {e}")
        sys.exit(1)

    if not quiet:
        cols_sorted = sorted(header_map.items(), key=lambda kv: kv[1])
        col_disp = ", ".join([f"{name}({get_column_letter(idx)})" for name, idx in cols_sorted])
        print(f"ヘッダー: {col_disp}")
        print("JSON処理完了")


if __name__ == "__main__":
    main()
