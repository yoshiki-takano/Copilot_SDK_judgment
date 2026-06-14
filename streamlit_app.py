from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import tkinter as tk
    from tkinter import filedialog
except Exception:
    tk = None
    filedialog = None

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MODEL_CANDIDATES = [
    "gpt-5.4",
    "gpt-5-mini",
    "gpt-5.3-codex",
    "claude-sonnet-4.6",
    "claude-sonnet-4.5",
    "claude-haiku-4.5",
    "claude-opus-4.6",
    "claude-opus-4.5",
]

SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

STREAMLIT_SECRET_TOKEN_KEYS = (
    "GITHUB_COPILOT_TOKEN",
    "COPILOT_GITHUB_TOKEN",
    "GITHUB_TOKEN",
)

_TOKEN_KEY_PATTERN = re.compile(r"^(?:export\s+)?([A-Za-z_][A-Za-z0-9_]*)\s*[:=]\s*(.+)$")


def extract_token_value(raw_text: str) -> str:
    """Extract a usable token from plain text or KEY=VALUE style input."""
    if not raw_text:
        return ""

    text = raw_text.replace("\ufeff", "").strip()
    if not text:
        return ""

    fallback = ""
    for raw_line in text.replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        m = _TOKEN_KEY_PATTERN.match(line)
        if m:
            key = m.group(1).strip().upper()
            value = m.group(2).strip()
            if " #" in value:
                value = value.split(" #", 1)[0].rstrip()
            if value and value[0] in {'"', "'"} and value[-1:] == value[0]:
                value = value[1:-1].strip()
            if value.lower().startswith("bearer "):
                value = value[7:].strip()
            if key in STREAMLIT_SECRET_TOKEN_KEYS and value:
                return value
            if value and not fallback:
                fallback = value
            continue

        candidate = line
        if candidate.lower().startswith("bearer "):
            candidate = candidate[7:].strip()
        if candidate and not fallback:
            fallback = candidate

    return fallback.strip()

@dataclass
class RunConfig:
    workspace: Path
    mode: str
    python_exe: str
    base_excel: Path
    base_name: str
    sheet_name: str
    part_count: int
    input_columns: list[str]
    prompt1: str
    prompt2: str
    token_file: str
    output_root: str
    parts_dir: str
    metadata_path: str
    outdir: str
    model_screening: str
    model_extraction: str


def norm(path_text: str) -> str:
    return path_text.replace("\\", "/")


def is_streamlit_cloud() -> bool:
    cloud_flag = str(os.getenv("STREAMLIT_SHARING_MODE", "")).strip().lower()
    runtime = str(os.getenv("STREAMLIT_RUNTIME", "")).strip().lower()
    cwd_text = norm(str(Path.cwd()))
    return (
        cloud_flag in {"1", "true", "yes"}
        or runtime == "cloud"
        or cwd_text.startswith("/mount/src")
    )


def detect_default_python(workspace: Path) -> str:
    if is_streamlit_cloud():
        # On Streamlit Cloud, force the current interpreter used by the app process.
        return norm(str(Path(sys.executable)))

    current = Path(sys.executable)
    if current.exists():
        return norm(str(current.resolve()))

    candidates = [
        workspace / ".venv" / "Scripts" / "python.exe",
        workspace / "venv" / "Scripts" / "python.exe",
        workspace / ".venv" / "bin" / "python",
        workspace / "venv" / "bin" / "python",
    ]
    for candidate in candidates:
        if candidate.exists():
            return norm(str(candidate.resolve()))
    found = shutil.which("python")
    return norm(found) if found else "python"


def runtime_inputs_dir(workspace: Path) -> Path:
    d = workspace / ".streamlit_runtime" / "inputs"
    d.mkdir(parents=True, exist_ok=True)
    return d


def pick_directory_dialog(initial_dir: str) -> str:
    if tk is None or filedialog is None:
        raise RuntimeError("tkinter が利用できないため、ダイアログを開けません")

    root = tk.Tk()
    try:
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askdirectory(
            title="出力先フォルダを選択",
            initialdir=initial_dir or str(Path.cwd()),
        )
    finally:
        root.destroy()

    return norm(str(Path(selected).expanduser())) if selected else ""


def save_uploaded_file(workspace: Path, uploaded_file, target_name: str | None = None) -> Path:
    out_dir = runtime_inputs_dir(workspace)
    filename = target_name or uploaded_file.name
    safe_name = Path(filename).name
    save_path = out_dir / safe_name
    data = uploaded_file.getvalue()
    save_path.write_bytes(data)
    return save_path


def resolve_python_command(py_cmd: str) -> str:
    if is_streamlit_cloud():
        return norm(str(Path(sys.executable)))
    value = py_cmd.strip()
    return value or "python"


def _safe_get_streamlit_secret(key: str) -> str:
    """Return a secret value or empty string when secrets are not configured."""
    try:
        return str(st.secrets.get(key, ""))
    except Exception:
        return ""


def save_token_from_secrets(workspace: Path) -> str:
    for key in STREAMLIT_SECRET_TOKEN_KEYS:
        token = extract_token_value(_safe_get_streamlit_secret(key))
        if not token:
            continue
        out_dir = runtime_inputs_dir(workspace)
        save_path = out_dir / "github_copilot_token_from_secrets.txt"
        save_path.write_text(token + "\n", encoding="utf-8")
        return norm(str(save_path))
    return ""


def resolve_token_file(workspace: Path) -> str:
    uploaded = st.session_state.get("_uploaded_token_path", "").strip()
    if uploaded and Path(uploaded).exists():
        return uploaded
    secret_token_path = save_token_from_secrets(workspace)
    if secret_token_path:
        st.session_state["_uploaded_token_path"] = secret_token_path
        return secret_token_path
    return ""


def fetch_copilot_status(py_cmd: str) -> dict[str, Any]:
    """Check SDK, CLI, and authentication status in one call."""
    script = r'''
import json
import os
import shutil
import subprocess
from pathlib import Path
from importlib import metadata

result_data = {}

# 1. Check SDK
try:
    import copilot
    try:
        from copilot._sdk_protocol_version import get_sdk_protocol_version
    except Exception:
        from copilot.sdk_protocol_version import get_sdk_protocol_version
    try:
        package_version = metadata.version("github-copilot-sdk")
    except Exception:
        package_version = getattr(copilot, "__version__", "unknown")
    result_data["sdk"] = {
        "version": package_version,
        "protocol": get_sdk_protocol_version(),
    }
except Exception as exc:
    result_data["sdk"] = {"error": str(exc)}

# 2. Check CLI Path and Version
def find_copilot_cli_path():
    for name in ("copilot", "copilot.exe"):
        found = shutil.which(name)
        if found:
            return found
    if os.name == "nt":
        winget_root = Path.home() / "AppData" / "Local" / "Microsoft" / "WinGet" / "Packages"
        if winget_root.exists():
            for candidate in winget_root.glob("GitHub.Copilot*/*copilot.exe"):
                if candidate.is_file():
                    return str(candidate)
    return None

try:
    cli_path = find_copilot_cli_path()
    if not cli_path:
        result_data["cli"] = {"error": "Copilot CLI not found"}
    else:
        result_data["cli"] = {"path": cli_path}
        
        # Get CLI version
        try:
            ver_result = subprocess.run(
                [cli_path, "--version"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=10,
            )
            if ver_result.returncode == 0:
                result_data["cli"]["version"] = ver_result.stdout.strip()
        except Exception as e:
            result_data["cli"]["version_error"] = str(e)
        
except Exception as exc:
    result_data["cli"] = {"error": str(exc)}

print(json.dumps(result_data, ensure_ascii=False))
'''
    try:
        resolved_py = resolve_python_command(py_cmd)
        result = subprocess.run(
            [resolved_py, "-c", script],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        raw = result.stdout.strip()
        return json.loads(raw) if raw else {"error": "No output"}
    except Exception as exc:
        return {"error": str(exc)}


def fetch_copilot_models(py_cmd: str, token_path: str) -> list[str]:
    script = r'''
import asyncio
import json
import os
import shutil
import sys
from pathlib import Path

from copilot import CopilotClient


TOKEN_KEYS = {"GITHUB_COPILOT_TOKEN", "COPILOT_GITHUB_TOKEN", "GITHUB_TOKEN"}


def extract_token_value(raw_text: str) -> str:
    if not raw_text:
        return ""
    text = raw_text.replace("\ufeff", "").strip()
    if not text:
        return ""

    fallback = ""
    for raw_line in text.replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        parsed = None
        for sep in ("=", ":"):
            if sep in line:
                left, right = line.split(sep, 1)
                parsed = (left.strip(), right.strip())
                break

        if parsed is not None and parsed[0]:
            key = parsed[0]
            if key.lower().startswith("export "):
                key = key[7:].strip()
            value = parsed[1]
            if " #" in value:
                value = value.split(" #", 1)[0].rstrip()
            if value and value[0] in {'"', "'"} and value[-1:] == value[0]:
                value = value[1:-1].strip()
            if value.lower().startswith("bearer "):
                value = value[7:].strip()
            if key.upper() in TOKEN_KEYS and value:
                return value
            if value and not fallback:
                fallback = value
            continue

        candidate = line
        if candidate.lower().startswith("bearer "):
            candidate = candidate[7:].strip()
        if candidate and not fallback:
            fallback = candidate

    return fallback.strip()


def find_copilot_cli_path():
    for name in ("copilot", "copilot.exe"):
        found = shutil.which(name)
        if found:
            return found

    if os.name == "nt":
        winget_root = Path.home() / "AppData" / "Local" / "Microsoft" / "WinGet" / "Packages"
        if winget_root.exists():
            for candidate in winget_root.glob("GitHub.Copilot*/*copilot.exe"):
                if candidate.is_file():
                    return str(candidate)
    return None


async def main():
    client = None
    try:
        token_path = sys.argv[1] if len(sys.argv) > 1 else ""
        token = ""
        if token_path:
            p = Path(token_path)
            if p.exists():
                token = extract_token_value(p.read_text(encoding="utf-8", errors="replace"))

        try:
            from copilot import SubprocessConfig  # type: ignore
            client = CopilotClient(
                SubprocessConfig(
                    cli_path=find_copilot_cli_path(),
                    github_token=token or None,
                    use_logged_in_user=False if token else None,
                )
            )
        except Exception:
            client = CopilotClient(
                github_token=token or None,
                use_logged_in_user=False if token else None,
            )

        await client.start()
        models = await client.list_models()
        print(json.dumps({"models": [{"id": m.id, "name": m.name} for m in models]}, ensure_ascii=False))
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        sys.exit(2)
    finally:
        if client is not None:
            try:
                await client.stop()
            except Exception:
                pass


asyncio.run(main())
'''

    resolved_py = resolve_python_command(py_cmd)
    result = subprocess.run(
        [resolved_py, "-c", script, token_path],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
    )
    if result.returncode != 0 and not result.stdout.strip():
        err = result.stderr.strip() or f"exit code {result.returncode}"
        raise RuntimeError(err)

    payload_text = result.stdout.strip()
    if not payload_text:
        return []

    payload = json.loads(payload_text)
    if isinstance(payload, dict) and payload.get("error"):
        raise RuntimeError(str(payload["error"]))

    raw_models = payload.get("models") if isinstance(payload, dict) else payload
    if not isinstance(raw_models, list):
        return []

    ids: list[str] = []
    seen: set[str] = set()
    for item in raw_models:
        if not isinstance(item, dict):
            continue
        model_id = str(item.get("id") or "").strip()
        if not model_id or model_id in seen:
            continue
        seen.add(model_id)
        ids.append(model_id)
    return ids


def script_for_mode(mode: str) -> str:
    if mode == "both":
        return "main_parallel.py"
    if mode == "screening":
        return "step1only_main_parallel.py"
    if mode == "extraction":
        return "step2only_main_parallel.py"
    raise ValueError(f"Unsupported mode: {mode}")


def required_prompts(mode: str) -> set[str]:
    if mode == "both":
        return {"prompt1", "prompt2"}
    if mode == "screening":
        return {"prompt1"}
    if mode == "extraction":
        return {"prompt2"}
    return set()


def parse_columns(raw_text: str) -> list[str]:
    parts = []
    for chunk in raw_text.replace("\r", "\n").replace(",", "\n").split("\n"):
        value = chunk.strip()
        if value:
            parts.append(value)
    deduped: list[str] = []
    seen: set[str] = set()
    for item in parts:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


_INPUT_PLACEHOLDER_RE = re.compile(r"\[入力\s*(\d+)\]")
_INPUT_BRACE_RE = re.compile(r"\{\{[^}]+\}\}")


def _extract_input_mapping_from_prompt(prompt_text: str) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for raw_line in prompt_text.replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        for m in _INPUT_PLACEHOLDER_RE.finditer(line):
            idx = int(m.group(1))
            col = line[: m.start()].strip()
            col = col.strip(":：- ")
            if col:
                mapping[idx] = col
    return mapping


def _extract_brace_columns_from_prompt(prompt_text: str) -> list[str]:
    columns: list[str] = []
    for raw_line in prompt_text.replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        if _INPUT_BRACE_RE.search(line):
            col = line.split("{{", 1)[0].strip()
            col = col.strip(":：- ")
            if col:
                columns.append(col)
    return columns


def detect_input_columns_from_prompts(mode: str, prompt1_path: str, prompt2_path: str) -> list[str]:
    needed = required_prompts(mode)
    prompt_paths: list[str] = []
    if "prompt1" in needed and prompt1_path:
        prompt_paths.append(prompt1_path)
    if "prompt2" in needed and prompt2_path:
        prompt_paths.append(prompt2_path)

    merged: dict[int, str] = {}
    brace_columns: list[str] = []
    for p in prompt_paths:
        text = Path(p).read_text(encoding="utf-8", errors="replace")
        for idx, col in _extract_input_mapping_from_prompt(text).items():
            merged[idx] = col
        brace_columns.extend(_extract_brace_columns_from_prompt(text))

    if not merged and not brace_columns:
        return []

    ordered = [merged[k] for k in sorted(merged.keys())]
    ordered.extend(brace_columns)
    deduped: list[str] = []
    seen: set[str] = set()
    for col in ordered:
        if col in seen:
            continue
        seen.add(col)
        deduped.append(col)
    return deduped


def _normalize_column_name(name: str) -> str:
    # Normalize punctuation/spacing variants so prompt labels can match Excel headers.
    text = name.strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s\u3000]+", "", text)
    text = re.sub(r"[-‐‑‒–—―ー_()]", "", text)
    return text


def resolve_detected_columns_to_headers(detected_columns: list[str], headers: list[str]) -> list[str]:
    norm_to_headers: dict[str, list[str]] = {}
    for h in headers:
        k = _normalize_column_name(h)
        norm_to_headers.setdefault(k, []).append(h)

    resolved: list[str] = []
    for col in detected_columns:
        if col in headers:
            resolved.append(col)
            continue

        candidates = norm_to_headers.get(_normalize_column_name(col), [])
        if len(candidates) == 1:
            resolved.append(candidates[0])
        else:
            # Keep original when unresolved/ambiguous; validation will show a clear error.
            resolved.append(col)
    return resolved


def validate_excel_path(src_path: Path) -> None:
    if not src_path.exists():
        raise ValueError(f"Excel file not found: {src_path}")
    if src_path.is_dir():
        raise ValueError(f"Base Excel path is a folder, not a file: {src_path}")
    if not src_path.is_file():
        raise ValueError(f"Base Excel path is not a file: {src_path}")
    if src_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        allowed = ", ".join(sorted(SUPPORTED_EXCEL_SUFFIXES))
        raise ValueError(f"Unsupported Excel format: {src_path.suffix}. Supported: {allowed}")


def read_excel_headers(src_path: Path, sheet_name: str | None) -> tuple[str, list[str]]:
    validate_excel_path(src_path)
    try:
        wb = load_workbook(filename=str(src_path), data_only=True, read_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            "Excel形式が不正です。Excelで開ける .xlsx/.xlsm/.xltx/.xltm を指定してください。"
        ) from exc
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers: list[str] = []
        for value in first_row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                headers.append(text)
        return ws.title, headers
    finally:
        wb.close()


def read_excel_sheet_names(src_path: Path) -> list[str]:
    validate_excel_path(src_path)
    try:
        wb = load_workbook(filename=str(src_path), data_only=True, read_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            "Excel形式が不正です。Excelで開ける .xlsx/.xlsm/.xltx/.xltm を指定してください。"
        ) from exc
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def split_excel(src_path: Path, parts_dir: Path, sheet_name: str | None, part_count: int) -> tuple[str, int, str]:
    validate_excel_path(src_path)
    if src_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("Input must be .xlsx or .xlsm for split output compatibility")
    if part_count <= 0:
        raise ValueError("part_count must be > 0")

    try:
        wb = load_workbook(filename=str(src_path), data_only=True, read_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            "Excel形式が不正です。Excelで開ける .xlsx/.xlsm を指定してください。"
        ) from exc
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        used_sheet_name = ws.title
        last_data_row = 1
        for row_idx, (value,) in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
            if value is not None and str(value) != "":
                last_data_row = row_idx

        if last_data_row < 2:
            return (src_path.stem, 0, used_sheet_name)

        data_rows = last_data_row - 1
        chunk_size = max(1, data_rows // part_count)
        parts_dir.mkdir(parents=True, exist_ok=True)
        base_name = src_path.stem

        header = None
        part_idx = 0
        rows_in_part = 0
        rows_target = 0
        new_wb = None
        new_ws = None
        made = 0

        def start_part(idx: int, target: int):
            wb_out = Workbook(write_only=True)
            ws_out = wb_out.create_sheet(title=used_sheet_name)
            ws_out.append(list(header) if header is not None else [])
            return wb_out, ws_out, target

        def finish_part(idx: int, wb_out):
            save_path = parts_dir / f"{base_name}_Part{idx}.xlsx"
            wb_out.save(str(save_path))
            wb_out.close()

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=last_data_row, values_only=True), start=1):
            if row_idx == 1:
                header = row
                continue

            if new_wb is None:
                part_idx += 1
                rows_target = chunk_size if part_idx < part_count else data_rows - made
                new_wb, new_ws, rows_target = start_part(part_idx, max(1, rows_target))
                rows_in_part = 0

            new_ws.append(list(row) if row is not None else [])
            rows_in_part += 1
            made += 1

            if rows_in_part >= rows_target:
                finish_part(part_idx, new_wb)
                new_wb = None
                new_ws = None

        if new_wb is not None:
            finish_part(part_idx, new_wb)

        return (base_name, part_idx, used_sheet_name)
    finally:
        wb.close()


def build_part_command(cfg: RunConfig, part_number: int) -> list[str]:
    script = str(cfg.workspace / script_for_mode(cfg.mode))
    input_path = str(Path(cfg.parts_dir) / f"{cfg.base_name}_Part{part_number}.xlsx")
    args = [
        cfg.python_exe,
        script,
        "--input",
        input_path,
        "--sheet",
        cfg.sheet_name,
    ]
    for col in cfg.input_columns:
        args += ["--input-column", col]

    needed = required_prompts(cfg.mode)
    if "prompt1" in needed:
        args += ["--prompt1", cfg.prompt1]
    if "prompt2" in needed:
        args += ["--prompt2", cfg.prompt2]

    args += [
        "--model-stage1",
        cfg.model_screening,
        "--model-stage2",
        cfg.model_extraction,
        "--outdir",
        cfg.outdir,
        "--no-ui",
    ]
    if cfg.token_file.strip():
        args += ["--apikey", cfg.token_file]
    return args


def build_merge_command(cfg: RunConfig, run_metadata_path: str) -> list[str]:
    script = str(cfg.workspace / script_for_mode(cfg.mode))
    merge_base = f"{cfg.base_name}_Part{cfg.part_count}"
    return [
        cfg.python_exe,
        script,
        "--merge-only",
        "--merge-base",
        merge_base,
        "--sheet",
        cfg.sheet_name,
        "--outdir",
        cfg.outdir,
        "--parts",
        str(cfg.part_count),
        "--no-ui",
        "--run-metadata",
        run_metadata_path,
    ]


def build_tasks_payload(cfg: RunConfig) -> dict[str, Any]:
    needed = required_prompts(cfg.mode)
    if "prompt1" in needed and not cfg.prompt1:
        raise ValueError("Prompt1 is required for this mode")
    if "prompt2" in needed and not cfg.prompt2:
        raise ValueError("Prompt2 is required for this mode")
    if cfg.part_count <= 0:
        raise ValueError("No split parts found")
    if not cfg.input_columns:
        raise ValueError("At least one input column is required")

    script = str(cfg.workspace / script_for_mode(cfg.mode))
    tasks: list[dict[str, Any]] = []
    labels: list[str] = []

    for n in range(1, cfg.part_count + 1):
        label = f"part{n}"
        labels.append(label)
        input_full = str(Path(cfg.parts_dir) / f"{cfg.base_name}_Part{n}.xlsx")

        args = [script, "--input", input_full, "--sheet", cfg.sheet_name]
        for col in cfg.input_columns:
            args += ["--input-column", col]

        if "prompt1" in needed:
            args += ["--prompt1", cfg.prompt1]
        if "prompt2" in needed:
            args += ["--prompt2", cfg.prompt2]

        args += [
            "--model-stage1",
            cfg.model_screening,
            "--model-stage2",
            cfg.model_extraction,
            "--outdir",
            cfg.outdir,
            "--no-ui",
        ]
        if cfg.token_file.strip():
            args += ["--apikey", cfg.token_file]

        tasks.append(
            {
                "label": label,
                "type": "shell",
                "command": cfg.python_exe,
                "args": args,
                "options": {"cwd": str(cfg.workspace)},
                "problemMatcher": [],
            }
        )

    merge_base = f"{cfg.base_name}_Part{cfg.part_count}"
    merge_args = [
        script,
        "--merge-only",
        "--merge-base",
        merge_base,
        "--sheet",
        cfg.sheet_name,
        "--outdir",
        cfg.outdir,
        "--parts",
        str(cfg.part_count),
        "--no-ui",
        "--run-metadata",
        cfg.metadata_path,
    ]
    tasks.append(
        {
            "label": "merge-only",
            "type": "shell",
            "command": cfg.python_exe,
            "args": merge_args,
            "options": {"cwd": str(cfg.workspace)},
            "problemMatcher": [],
        }
    )

    tasks.append({"label": "run-all-parallel", "dependsOn": labels})
    tasks.append(
        {
            "label": "run-all-parallel-then-merge",
            "dependsOn": ["run-all-parallel", "merge-only"],
            "dependsOrder": "sequence",
        }
    )

    return {"version": "2.0.0", "tasks": tasks}


def write_tasks_json(cfg: RunConfig, tasks_path: Path) -> None:
    tasks_path.parent.mkdir(parents=True, exist_ok=True)
    payload = build_tasks_payload(cfg)
    tasks_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


_ANSI_ESCAPE = re.compile(r"\x1b\[[0-9;]*[mGKHF]")


def _decode_output(raw: bytes) -> str:
    """Decode subprocess bytes with utf-8 first, then cp932 fallback."""
    for enc in ("utf-8", "cp932", "utf-8-sig"):
        try:
            text = raw.decode(enc)
            return _ANSI_ESCAPE.sub("", text)
        except (UnicodeDecodeError, LookupError):
            continue
    return _ANSI_ESCAPE.sub("", raw.decode("utf-8", errors="replace"))


def run_command(args: list[str], cwd: Path) -> dict[str, Any]:
    started = datetime.now()
    proc = subprocess.run(
        args,
        cwd=str(cwd),
        capture_output=True,
    )
    finished = datetime.now()
    return {
        "args": args,
        "returncode": proc.returncode,
        "stdout": _decode_output(proc.stdout),
        "stderr": _decode_output(proc.stderr),
        "started_at": started.isoformat(timespec="seconds"),
        "finished_at": finished.isoformat(timespec="seconds"),
    }


def write_run_metadata(cfg: RunConfig, metadata_path: Path) -> None:
    metadata_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "mode": cfg.mode,
        "parts": cfg.part_count,
        "base_excel_path": norm(str(cfg.base_excel)),
        "base_name": cfg.base_name,
        "sheet_name": cfg.sheet_name,
        "input_columns": cfg.input_columns,
        "prompt1_path": norm(cfg.prompt1),
        "prompt2_path": norm(cfg.prompt2),
        "apikey_path": norm(cfg.token_file),
        "out_dir": norm(cfg.outdir),
        "model_screening": cfg.model_screening,
        "model_extraction": cfg.model_extraction,
        # keep legacy keys for merge/read compatibility with older artifacts
        "model_stage1": cfg.model_screening,
        "model_stage2": cfg.model_extraction,
        "python_command": cfg.python_exe,
    }
    metadata_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def validate_config(cfg: RunConfig) -> list[str]:
    errs: list[str] = []
    if not cfg.workspace.exists():
        errs.append("workspace が存在しません")
    if str(cfg.base_excel).strip() in {"", "."}:
        errs.append("分析対象の Excel をアップロードしてください")
    if not cfg.base_excel.exists():
        errs.append("分析対象のExcelが存在しません")
    elif not cfg.base_excel.is_file():
        errs.append("分析対象のExcelはファイルを指定してください（フォルダ不可）")
    if cfg.python_exe != "python" and not Path(cfg.python_exe).exists():
        # Allow command-style python values (python3, python3.12, etc.) when resolvable in PATH.
        if shutil.which(cfg.python_exe) is None:
            errs.append("Python 実行ファイルが存在しません")
    if not cfg.output_root.strip():
        errs.append("出力フォルダを指定してください")
    else:
        try:
            Path(cfg.output_root).mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            errs.append(f"出力フォルダを作成できません: {exc}")
    if cfg.token_file.strip() and not Path(cfg.token_file).exists():
        errs.append("指定されたToken file が存在しません")
    if "prompt1" in required_prompts(cfg.mode) and not Path(cfg.prompt1).exists():
        errs.append("Prompt1 をアップロードしてください")
    if "prompt2" in required_prompts(cfg.mode) and not Path(cfg.prompt2).exists():
        errs.append("Prompt2 をアップロードしてください")
    if not cfg.input_columns:
        errs.append("Prompt から入力列を特定できません")
    elif cfg.base_excel.exists() and cfg.base_excel.is_file():
        try:
            _, headers = read_excel_headers(cfg.base_excel, cfg.sheet_name)
            missing = [col for col in cfg.input_columns if col not in headers]
            if missing:
                errs.append(f"入力ファイルに対象列がありません: {', '.join(missing)}")
        except Exception as exc:
            errs.append(f"Excelヘッダ確認に失敗しました: {exc}")
    return errs


def get_config_from_ui() -> RunConfig:
    workspace = Path(st.session_state.workspace).expanduser().resolve()
    base_excel_text = st.session_state.get("_uploaded_base_excel_path", "").strip()
    base_excel = Path(base_excel_text).expanduser() if base_excel_text else Path()
    python_exe = resolve_python_command(st.session_state.python_exe)
    prompt1 = st.session_state.get("_uploaded_prompt1_path", "").strip()
    prompt2 = st.session_state.get("_uploaded_prompt2_path", "").strip()
    columns = detect_input_columns_from_prompts(st.session_state.mode, prompt1, prompt2)
    # base_name and sheet_name are derived/updated via backing keys (_base_name, _sheet_name)
    base_name = st.session_state.get("_base_name", "").strip() or base_excel.stem
    sheet_name = st.session_state.get("_sheet_name", "Sheet1").strip() or "Sheet1"
    output_root = st.session_state.get("_output_root", norm(str(workspace / "out"))).strip()
    output_root = norm(str(Path(output_root).expanduser().resolve()))
    parts_dir = norm(str(Path(output_root) / "parts"))
    outputs_dir = norm(str(Path(output_root) / "outputs"))
    metadata_path = norm(str(Path(output_root) / ".vscode" / "copilot_run_metadata.json"))

    if columns and base_excel.exists() and base_excel.is_file():
        try:
            _, headers = read_excel_headers(base_excel, sheet_name)
            columns = resolve_detected_columns_to_headers(columns, headers)
        except Exception:
            pass

    token_file = resolve_token_file(workspace)

    return RunConfig(
        workspace=workspace,
        mode=st.session_state.mode,
        python_exe=python_exe,
        base_excel=base_excel,
        base_name=base_name,
        sheet_name=sheet_name,
        part_count=int(st.session_state.get("_part_count", 10)),
        input_columns=columns,
        prompt1=prompt1,
        prompt2=prompt2,
        token_file=token_file,
        output_root=output_root,
        parts_dir=parts_dir,
        metadata_path=metadata_path,
        outdir=outputs_dir,
        model_screening=st.session_state.model_screening,
        model_extraction=st.session_state.model_extraction,
    )


def init_state() -> None:
    ws = Path.cwd().resolve()
    default_output_root = norm(str(ws / "out"))
    defaults = {
        "workspace": str(ws),
        "python_exe": detect_default_python(ws),
        "mode": "screening",
        # backing keys — not bound to widgets, safe to update programmatically
        "_base_name": "",
        "_sheet_name": "Sheet1",
        "_part_count": 10,
        "_uploaded_base_excel_path": "",
        "_uploaded_prompt1_path": "",
        "_uploaded_prompt2_path": "",
        "_uploaded_token_path": "",
        "_output_root": default_output_root,
        "_model_options": list(MODEL_CANDIDATES),
        "_model_status": "静的候補を使用中",
        "_model_refresh_notice": "",
        "model_screening": MODEL_CANDIDATES[0],
        "model_extraction": MODEL_CANDIDATES[0],
        "split_info": "未実行",
        "_last_split_parts": [],
        "last_logs": {},
        "_is_cloud": is_streamlit_cloud(),
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    # Streamlit Cloud can retain older session_state values across reruns.
    # Force command-style python so subprocess checks run in the app venv.
    if st.session_state.get("_is_cloud", False):
        st.session_state["python_exe"] = resolve_python_command(st.session_state.get("python_exe", "python"))

    # In Streamlit Cloud, prefer token from st.secrets when available.
    if not st.session_state.get("_uploaded_token_path"):
        token_path = save_token_from_secrets(ws)
        if token_path:
            st.session_state["_uploaded_token_path"] = token_path

    # Migrate legacy keys if user session has old names.
    if "model_stage1" in st.session_state and "model_screening" not in st.session_state:
        st.session_state["model_screening"] = st.session_state["model_stage1"]
    if "model_stage2" in st.session_state and "model_extraction" not in st.session_state:
        st.session_state["model_extraction"] = st.session_state["model_stage2"]


def main() -> None:
    st.set_page_config(page_title="Copilot SDK Runner", layout="wide")
    init_state()
    is_cloud = bool(st.session_state.get("_is_cloud", False))

    st.title("Copilot SDK Streamlit Runner")
    if is_cloud:
        st.caption("Streamlit Cloud モード: ヘッドレス実行 + Secrets ベースの認証を推奨します")
    else:
        st.caption("ローカル環境でCopilot SDK/CLIを実行します")

    tab_wsp, tab_cfg = st.tabs(["Workspace", "Execute"])
#    tab_wsp, tab_cfg, tab_exec = st.tabs(["Workspace", "Input files setting", "Execute"])

    with tab_wsp:
        st.subheader("ワークスペース設定")
        ws_dir = Path(st.session_state.workspace).expanduser().resolve()

        st.markdown("**実行環境**")
        st.caption("プログラムファイルの保存されているフォルダが自動的に選択されます")
        e1, e2 = st.columns(2)
        with e1:
            st.caption(f"Workspace: {norm(str(ws_dir))}")
        with e2:
            st.caption(f"Python: {st.session_state.python_exe}")

        st.markdown("**出力フォルダ**")
        st.caption("生成物は指定したフォルダ内に保存されます")

        def _sync_output_root():
            text = st.session_state.get("_output_root_widget", "").strip()
            if text:
                st.session_state["_output_root"] = norm(str(Path(text).expanduser()))

        out_text_col, out_btn_col = st.columns([6, 1])
        with out_text_col:
            st.text_input(
                "Output root folder",
                value=st.session_state.get("_output_root", norm(str(ws_dir / "out"))),
                key="_output_root_widget",
                on_change=_sync_output_root,
                help="生成物は <出力先>/.vscode/copilot_run_metadata.json, <出力先>/parts, <出力先>/outputs に保存されます",
            )
        with out_btn_col:
            st.markdown("<div style='height: 1.75rem;'></div>", unsafe_allow_html=True)
            if st.button("参照", key="_pick_output_root", disabled=is_cloud or tk is None or filedialog is None):
                try:
                    selected = pick_directory_dialog(st.session_state.get("_output_root", norm(str(ws_dir / "out"))))
                    if selected:
                        st.session_state["_output_root"] = selected
                        st.rerun()
                except Exception as exc:
                    st.warning(f"フォルダ選択ダイアログを開けませんでした: {exc}")
        if is_cloud:
            st.caption("Cloudではフォルダ参照ダイアログを利用できないため、パスを直接入力してください")

        output_root_preview = st.session_state.get("_output_root", norm(str(ws_dir / "out")))
        st.caption(f"Metadata: {norm(str(Path(output_root_preview) / '.vscode' / 'copilot_run_metadata.json'))}")
        st.caption(f"Parts: {norm(str(Path(output_root_preview) / 'parts'))}")
        st.caption(f"Outputs: {norm(str(Path(output_root_preview) / 'outputs'))}")

        uploaded_token = st.file_uploader("Token file upload", type=["txt"], key="_upload_token")
        if uploaded_token is not None:
            saved = save_uploaded_file(ws_dir, uploaded_token, "github_copilot_token.txt")
            st.session_state["_uploaded_token_path"] = norm(str(saved))
        token_path = resolve_token_file(ws_dir)
        if token_path:
            if "from_secrets" in token_path:
                st.success("Copilot token: Streamlit Secrets から読み込み済み")
            else:
                st.caption(f"Token file: {token_path}")
        else:
            st.caption("Token未設定時は CLI ログインユーザーを利用します（Cloudでは非推奨）")

        st.markdown("**Copilot SDK・CLI状態確認**")
        if st.button("統合状態を確認"):
            with st.spinner("状態確認中..."):
                status = fetch_copilot_status(st.session_state.python_exe.strip() or "python")
            
            # SDK status
            if status.get("sdk", {}).get("error"):
                st.error(f"SDK: {status['sdk']['error']}")
            else:
                sdk_version = status.get("sdk", {}).get("version", "unknown")
                sdk_protocol = status.get("sdk", {}).get("protocol", "unknown")
                min_protocol = 3
                if int(sdk_protocol or 0) < min_protocol:
                    st.warning(f"SDK: 更新が必要です (version={sdk_version}, protocol={sdk_protocol}, 要件: >= {min_protocol})")
                else:
                    st.success(f"SDK OK: version={sdk_version}, protocol={sdk_protocol}")
            
            st.divider()
            
            # CLI status
            cli_info = status.get("cli", {})
            if cli_info.get("error"):
                if is_cloud:
                    pass
                else:
                    st.error(f"CLI: {cli_info['error']}")
            else:
                cli_path = cli_info.get("path", "unknown")
                cli_version = cli_info.get("version", "取得できません")
                st.success(f"CLI OK\n- Path: {cli_path}\n- Version: {cli_version}")

    with tab_cfg:
        st.subheader("実行設定")

        st.markdown("**入力データ**")
        _sh = st.session_state.get("_sheet_name", "Sheet1")

        u1 = st.columns([1])[0]
        with u1:
            uploaded_excel = st.file_uploader(
                "分析対象Excel upload",
                type=["xlsx", "xlsm"],
                key="_upload_base_excel",
            )
            if uploaded_excel is not None:
                saved = save_uploaded_file(ws_dir, uploaded_excel)
                st.session_state["_uploaded_base_excel_path"] = norm(str(saved))

        sheet_names: list[str] = []
        sheet_err = ""
        base_excel_path = st.session_state.get("_uploaded_base_excel_path", "").strip()
        if base_excel_path:
            try:
                sheet_names = read_excel_sheet_names(Path(base_excel_path))
            except Exception as exc:
                sheet_err = str(exc)

        if sheet_err:
            st.warning(sheet_err)
        elif len(sheet_names) <= 1:
            if len(sheet_names) == 1:
                st.session_state["_sheet_name"] = sheet_names[0]
                st.caption(f"Sheet: {sheet_names[0]}（1シートのため自動固定）")
            else:
                st.caption("Sheet: Base Excel アップロード後に自動設定")
        else:
            if _sh not in sheet_names:
                st.session_state["_sheet_name"] = sheet_names[0]

            def _sync_sheet_name():
                st.session_state["_sheet_name"] = st.session_state["_sheet_name_widget"]

            st.selectbox(
                "Sheet",
                options=sheet_names,
                index=sheet_names.index(st.session_state.get("_sheet_name", sheet_names[0])),
                key="_sheet_name_widget",
                on_change=_sync_sheet_name,
            )

        st.markdown("**分析設定**")
        st.caption("stage 1: screening (ノイズ落とし), stage 2: extraction (stage 1の該当に対して要素抽出)")
        mode_cols = st.columns([1, 4])
        with mode_cols[0]:
            st.selectbox(
                "Mode",
                ["screening", "extraction", "both"],
                key="mode",
                format_func=lambda m: {
                    "screening": "stage 1",
                    "extraction": "stage 2",
                    "both": "stage 1 + stage 2",
                }.get(m, m),
            )

        mode = st.session_state.mode

        st.markdown("**プロンプト**")
        p1, p2 = st.columns(2)
        with p1:
            uploaded_prompt1 = st.file_uploader("Stage 1 prompt upload", type=["txt"], key="_upload_prompt1")
            if uploaded_prompt1 is not None:
                saved = save_uploaded_file(ws_dir, uploaded_prompt1)
                st.session_state["_uploaded_prompt1_path"] = norm(str(saved))
            if not st.session_state.get("_uploaded_prompt1_path") and mode in {"screening", "both"}:
                st.warning("Stage 1 prompt が必要です")
        with p2:
            uploaded_prompt2 = st.file_uploader("Stage 2 prompt upload", type=["txt"], key="_upload_prompt2")
            if uploaded_prompt2 is not None:
                saved = save_uploaded_file(ws_dir, uploaded_prompt2)
                st.session_state["_uploaded_prompt2_path"] = norm(str(saved))
            if not st.session_state.get("_uploaded_prompt2_path") and mode in {"extraction", "both"}:
                st.warning("Stage 2 prompt が必要です")

        st.markdown("**モデル選択**")
        notice = st.session_state.get("_model_refresh_notice", "")
        if notice:
            st.success(notice)
            st.session_state["_model_refresh_notice"] = ""
        model_options = st.session_state.get("_model_options", MODEL_CANDIDATES)
        m2, m3, m_btn = st.columns([2, 2, 1])
        model_options = st.session_state.get("_model_options", MODEL_CANDIDATES)
        if mode == "screening":
            with m2:
                st.selectbox("Model for Stage 1", model_options, key="model_screening")
        elif mode == "extraction":
            with m2:
                st.selectbox("Model for Stage 2", model_options, key="model_extraction")
        else:
            with m2:
                st.selectbox("Model for Stage 1", model_options, key="model_screening")
            with m3:
                st.selectbox("Model for Stage 2", model_options, key="model_extraction")
        with m_btn:
            st.markdown("<div style='height: 1.75rem;'></div>", unsafe_allow_html=True)
            if st.button("モデル再取得"):
                try:
                    token_path = resolve_token_file(ws_dir)
                    fetched = fetch_copilot_models(st.session_state.python_exe.strip() or "python", token_path)
                    if fetched:
                        st.session_state["_model_options"] = fetched
                        model_options = fetched
                        if st.session_state.model_screening not in model_options:
                            st.session_state.model_screening = model_options[0]
                        if st.session_state.model_extraction not in model_options:
                            st.session_state.model_extraction = model_options[0]
                        st.session_state["_model_status"] = f"動的取得: {len(fetched)}件"
                        st.session_state["_model_refresh_notice"] = f"モデル再取得に成功しました（{len(fetched)}件）"
                        st.rerun()
                    else:
                        st.session_state["_model_status"] = "動的取得結果が空のため静的候補を継続"
                except Exception as exc:
                    st.session_state["_model_options"] = list(MODEL_CANDIDATES)
                    model_options = list(MODEL_CANDIDATES)
                    st.session_state["_model_status"] = f"動的取得失敗: {exc}"
        st.caption(st.session_state.get("_model_status", ""))

        st.markdown("**自動入力列**")
        detected_cols = detect_input_columns_from_prompts(
            st.session_state.mode,
            st.session_state.get("_uploaded_prompt1_path", "").strip(),
            st.session_state.get("_uploaded_prompt2_path", "").strip(),
        )
        st.caption("Input columns は Prompt の [入力n] から自動決定されます")
        if detected_cols:
            st.caption("Auto input columns: " + ", ".join(detected_cols))
        else:
            st.warning("Prompt から入力列を抽出できません")

        st.markdown("**チェック**")
        run_preflight = st.button("Preflight check")

        if run_preflight:
            cfg = get_config_from_ui()
            errors = validate_config(cfg)
            if errors:
                st.error("\n".join(f"- {item}" for item in errors))
            else:
                used_sheet, headers = read_excel_headers(cfg.base_excel, cfg.sheet_name)
                st.session_state["_sheet_name"] = used_sheet
                st.session_state["_base_name"] = cfg.base_excel.stem
                st.success("設定チェックOK")
                st.caption(f"ヘッダ確認: シート={used_sheet}, 列数={len(headers)}")
                preview = build_part_command(cfg, 1)
                st.code(" ".join(preview), language="bash")

    # with tab_exec:
        st.subheader("実行")
        cfg = get_config_from_ui()
        stop_on_error = st.checkbox("逐次実行時にエラーで停止", value=True)
        max_workers = st.slider("Parallel workers", min_value=1, max_value=10, value=10)
        st.caption("Parallel workers 数は、同時に実行する数を表し、Run selected part / Run all parts の実行前にExcelの自動分割に使用されます")

        st.markdown("Part to run (分割したファイルの一部を実行)")
        part_col, run_col, _ = st.columns([1.25, 1.15, 3.6])
        with part_col:
            selected_part = st.selectbox(
                "Part to run",
                options=list(range(1, 11)),
                index=0,
                label_visibility="collapsed",
            )
        with run_col:
            run_selected = st.button("Run selected part", use_container_width=True)

        def _prepare_split_for_execution(current_cfg: RunConfig, split_parts: int) -> tuple[RunConfig, list[int]]:
            parts_dir = Path(current_cfg.parts_dir)
            parts_dir.mkdir(parents=True, exist_ok=True)
            base_name, made, used_sheet = split_excel(
                current_cfg.base_excel,
                parts_dir,
                current_cfg.sheet_name,
                split_parts,
            )
            current_cfg.base_name = base_name
            current_cfg.sheet_name = used_sheet
            current_cfg.part_count = made
            st.session_state["_base_name"] = base_name
            st.session_state["_sheet_name"] = used_sheet
            st.session_state["_part_count"] = made
            st.session_state["_last_split_parts"] = list(range(1, made + 1))
            st.session_state.split_info = f"完了: {made} parts"
            return current_cfg, st.session_state["_last_split_parts"]

        part_all_col, _ = st.columns([2, 2])
        with part_all_col:
            st.markdown("Run all parts (分割したファイルの全てを実行し結合する)", unsafe_allow_html=True)
            run_all = st.button("Run all parts ", use_container_width=True)
        
        # Streamlit Cloud での並列実行は不安定なため、Cloud では sequential 実行を強制
        is_cloud = bool(st.session_state.get("_is_cloud", False))
        if is_cloud and run_all:
            st.info("⚠️ Streamlit Cloud では sequential 実行を使用します（スレッドセーフ）")

        execution_area = st.container()

        action = ""
        if run_selected:
            action = "selected"
        elif run_all:
            action = "all"

        if action:
            with execution_area:
                errors = validate_config(cfg)
                if errors:
                    st.error("\n".join(f"- {item}" for item in errors))
                else:
                    try:
                        cfg, part_numbers = _prepare_split_for_execution(cfg, max_workers)
                        st.info(f"Split完了: {len(part_numbers)} parts")
                    except Exception as exc:
                        st.error(f"Split失敗: {exc}")
                        part_numbers = []

                    if not part_numbers:
                        st.error("実行対象のPartがありません")
                    elif action == "selected":
                        if selected_part not in part_numbers:
                            st.error(f"選択したPart{selected_part}は存在しません。利用可能: {part_numbers}")
                        else:
                            Path(cfg.outdir).mkdir(parents=True, exist_ok=True)
                            write_run_metadata(cfg, Path(cfg.metadata_path))
                            write_tasks_json(cfg, Path(cfg.metadata_path).parent / "tasks.json")
                            progress = st.progress(0)
                            status = st.empty()
                            status.write(f"part{selected_part}: 実行中...")
                            result = run_command(build_part_command(cfg, selected_part), cfg.workspace)
                            progress.progress(1.0)
                            status.write(f"part{selected_part}: rc={result['returncode']}")
                            st.session_state.last_logs = {f"part{selected_part}": result}
                            if result["returncode"] == 0:
                                st.success(f"part{selected_part} 成功")
                                
                                # Part ファイルをダウンロード可能にする
                                outdir_path = Path(cfg.outdir)
                                part_files = list(outdir_path.glob(f"*_Part{selected_part}.xlsx"))
                                if part_files:
                                    part_file = sorted(part_files, key=lambda p: p.stat().st_mtime)[-1]
                                    with open(part_file, "rb") as f:
                                        file_data = f.read()
                                    st.download_button(
                                        label=f"📥 {part_file.name} をダウンロード",
                                        data=file_data,
                                        file_name=part_file.name,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.error(f"part{selected_part} 失敗")
                    else:
                        Path(cfg.outdir).mkdir(parents=True, exist_ok=True)
                        write_run_metadata(cfg, Path(cfg.metadata_path))
                        write_tasks_json(cfg, Path(cfg.metadata_path).parent / "tasks.json")
                        results: dict[str, Any] = {}
                        progress = st.progress(0)
                        status = st.empty()
                        done = 0
                        status_history: list[str] = []

                        def push_status(msg: str) -> None:
                            status_history.append(msg)
                            # Keep the latest lines visible while processing many parts.
                            status.markdown("  \n".join(status_history[-30:]))

                        # Streamlit Cloud では parallel execution（ThreadPoolExecutor）が不安定なため sequential に固定
                        force_sequential = is_cloud or max_workers == 1
                        
                        if force_sequential:
                            for n in part_numbers:
                                push_status(f"part{n}: 実行中...")
                                result = run_command(build_part_command(cfg, n), cfg.workspace)
                                results[f"part{n}"] = result
                                done += 1
                                progress.progress(done / len(part_numbers))
                                push_status(f"part{n}: rc={result['returncode']}")
                                if stop_on_error and result["returncode"] != 0:
                                    break
                        else:
                            with ThreadPoolExecutor(max_workers=max_workers) as exe:
                                futures = {}
                                for n in part_numbers:
                                    push_status(f"part{n}: 実行中...")
                                    futures[exe.submit(run_command, build_part_command(cfg, n), cfg.workspace)] = n
                                for fut in as_completed(futures):
                                    n = futures[fut]
                                    result = fut.result()
                                    results[f"part{n}"] = result
                                    done += 1
                                    progress.progress(done / len(part_numbers))
                                    push_status(f"part{n}: rc={result['returncode']}")

                        failed = [k for k, v in results.items() if v["returncode"] != 0]
                        if failed:
                            st.error(f"失敗: {', '.join(sorted(failed))}")
                            st.warning("失敗PartがあるためMergeは実行しません")
                        else:
                            merge_result = run_command(build_merge_command(cfg, cfg.metadata_path), cfg.workspace)
                            results["merge"] = merge_result
                            if merge_result["returncode"] == 0:
                                st.success("全part成功 + merge成功")
                                
                                # MERGED ファイルをダウンロード可能にする
                                outdir_path = Path(cfg.outdir)
                                if not outdir_path.exists():
                                    st.warning(f"出力ディレクトリが見つかりません: {outdir_path}")
                                else:
                                    # MERGED ファイルを探す
                                    merged_files = list(outdir_path.glob("*MERGED.xlsx")) + list(outdir_path.glob("*merged.xlsx"))
                                    
                                    if not merged_files:
                                        # すべての xlsx ファイルをリストしてデバッグ
                                        all_files = sorted(list(outdir_path.glob("*.xlsx")), key=lambda p: p.stat().st_mtime, reverse=True)
                                        
                                        if not all_files:
                                            st.warning(f"出力ディレクトリにファイルがありません: {outdir_path}")
                                        else:
                                            st.info(f"出力ディレクトリ内のファイル: {[f.name for f in all_files[:20]]}")
                                            # 最新ファイルをダウンロード対象にする（Part ファイルではなく最後に生成されたもの）
                                            newest_file = all_files[0]
                                            st.write(f"最新ファイル: {newest_file.name}")
                                            with open(newest_file, "rb") as f:
                                                file_data = f.read()
                                            st.download_button(
                                                label=f"📥 {newest_file.name} をダウンロード",
                                                data=file_data,
                                                file_name=newest_file.name,
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                            )
                                    else:
                                        # MERGED ファイルが見つかった
                                        merged_file = sorted(merged_files, key=lambda p: p.stat().st_mtime)[-1]
                                        with open(merged_file, "rb") as f:
                                            file_data = f.read()
                                        st.download_button(
                                            label=f"📥 {merged_file.name} をダウンロード",
                                            data=file_data,
                                            file_name=merged_file.name,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )
                            else:
                                st.error("全part成功 / merge失敗")

                        st.session_state.last_logs = results

        if st.session_state.last_logs:
            with st.expander("Latest logs", expanded=False):
                for key in sorted(st.session_state.last_logs.keys()):
                    log = st.session_state.last_logs[key]
                    st.markdown(f"**{key}**")
                    st.write(f"returncode: {log['returncode']}")
                    st.code(log["stdout"] or "(no stdout)")
                    if log["stderr"]:
                        st.code(log["stderr"])


if __name__ == "__main__":
    main()
