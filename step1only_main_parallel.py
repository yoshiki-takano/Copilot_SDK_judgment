# 実行用プログラム (Google翻訳版) 並列起動対応・CLI引数対応版
# - 各 Part の処理（翻訳→Stage1→Stage2 書き込み）
# - 追加機能:
#     --merge-after : この実行の最後に Part1..N の最新出力を結合
#     --merge-only  : 処理を行わず結合だけを実行して終了（専用タスク用）
#     --merge-base  : 結合対象のベース名（拡張子なし・末尾が PartN）
#     --parts       : パート数（既定 10）

import argparse
import re
import pandas as pd
import sys
import subprocess
import hashlib
import tempfile
from datetime import datetime
from openpyxl import load_workbook, Workbook
import glob
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    simpledialog = None
from tqdm import tqdm
import os
import json
from openpyxl.utils import get_column_letter

# --- スクリプトの場所を基準にパスを設定 ---
script_dir = os.path.dirname(os.path.abspath(__file__))

# 参照する外部スクリプトのファイル名（相対はこのスクリプトの場所基準）
gemini_summarize_filename = "Copilot_Calling.py"
split_abstract_filename   = "rev_JSON_Fillout in Excel.py"


# 既定のAPIキーのファイルパス（引数で上書き可能）
default_api_key_path = ""

# --- 使用するAIモデル（引数で上書き可能な既定値） ---
default_model_stage1 = 'gpt-5'   # スクリーニング用モデル
default_model_stage2 = 'gpt-5'   # 詳細抽出用モデル

# 各ファイルの完全なパス
Gemini_summarize_path = os.path.join(script_dir, gemini_summarize_filename)
Split_Abstract_path   = os.path.join(script_dir, split_abstract_filename)

# -------------------------
# 引数パーサ
# -------------------------
def parse_args():
    p = argparse.ArgumentParser(description="Excel（分割済み）を1本処理。UI/非UI両対応。")
    p.add_argument("--input",   type=str, help="入力Excelのパス（指定時はUIスキップ）")
    p.add_argument("--sheet",   type=str, help="シート名（指定時はUIスキップ）")
    p.add_argument("--prompt1", type=str, help="1段階目（スクリーニング）プロンプトTXTのパス")
    p.add_argument("--prompt2", type=str, help="2段階目（詳細抽出）プロンプトTXTのパス")
    p.add_argument("--apikey", "--token", dest="apikey", type=str, default=default_api_key_path,
                   help="Copilot用GitHubトークンのパス")
    p.add_argument("--outdir",  type=str, help="出力先フォルダ（省略時は入力Excelと同じ場所）")
    p.add_argument("--no-ui", dest="no_ui", action="store_true", help="UIダイアログを出さない（完全ヘッドレス）")
    p.add_argument("--model-stage1", type=str, default=default_model_stage1,
                   help="1段階目で使うCopilotモデル名")
    p.add_argument("--model-stage2", type=str, default=default_model_stage2,
                   help="2段階目で使うCopilotモデル名")
    p.add_argument("--input-column", dest="input_columns", action="append",
                   help="プロンプトの [入力N] に差し込むExcel列名。複数指定可")
    p.add_argument("--web-search", dest="web_search", action="store_true",
                   help="互換オプション（Copilot SDKでは無視されます）")
    # ▼ 結合オプション
    p.add_argument("--merge-after", dest="merge_after", action="store_true",
                   help="この実行の最後に Part1〜PartN の最新出力を1つのExcelに結合する")
    p.add_argument("--parts", type=int, default=10, help="パート数（既定10）")
    p.add_argument("--merge-only", dest="merge_only", action="store_true",
                   help="処理は行わず、Part1〜PartN の最新出力を結合して終了する")
    p.add_argument("--merge-base", type=str,
                   help="結合のベース名（拡張子なし・末尾が PartN）。例: 標的_…_Part10")
    p.add_argument("--run-metadata", type=str,
                   help="MERGEDファイルの実行条件シートに記録するJSONファイルのパス")
    return p.parse_args()

args = parse_args()

DEFAULT_INPUT_COLUMNS = ['請求項（英語）', 'タイトル（英語）']

# ---- JSON 解析ヘルパー（Stage1 の厳格バリデーションに使用） ----
from typing import Any, Dict, List, Tuple, Union
Scalar = Union[str, int, float, bool, None]
JSONType = Union[Dict[str, Any], List[Any], Scalar]

def strip_code_fences(s: str) -> str:
    """ ```json ... ``` のようなコードフェンスを剥がす """
    if not isinstance(s, str):
        return s
    fence_re = re.compile(r"^\s*```(?:json|JSON)?\s*([\s\S]*?)\s*```\s*$")
    m = fence_re.match(s.strip())
    return m.group(1) if m else s

def parse_json_input(inp: str) -> JSONType:
    """
    文字列 inp を JSON として厳格にロード。
    失敗したら {'_raw_text': ..., '_json_error': ...} を返して非JSON扱いにする。
    """
    if os.path.isfile(inp):
        with open(inp, "r", encoding="utf-8") as f:
            raw = f.read()
    else:
        raw = inp
    raw = strip_code_fences(raw).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        return {"_raw_text": raw, "_json_error": str(e)}

# -------------------------
# UI使用可否と通知ヘルパ
# -------------------------
use_ui = not args.no_ui
if use_ui and (tk is None or filedialog is None or messagebox is None or simpledialog is None):
    print("[WARN] tkinter が利用できないため --no-ui モードで実行します", file=sys.stderr)
    use_ui = False

def notify_info(title, msg):
    if use_ui:
        try:
            messagebox.showinfo(title, msg)
        except Exception:
            print(f"[INFO] {title}: {msg}")
    else:
        print(f"[INFO] {title}: {msg}")

def notify_error(title, msg):
    if use_ui:
        try:
            messagebox.showerror(title, msg)
        except Exception:
            print(f"[ERROR] {title}: {msg}", file=sys.stderr)
    else:
        print(f"[ERROR] {title}: {msg}", file=sys.stderr)

# --- Tkinterルート（UIを使う場合のみ初期化） ---
if use_ui:
    root = tk.Tk()
    root.withdraw()
else:
    root = None

# -------------------------
# UI関数（必要な時だけ使用）
# -------------------------
def select_file(title, filetypes):
    if not use_ui:
        return None
    return filedialog.askopenfilename(title=title, filetypes=filetypes)

def ask_sheetname(default_name="savedrecs"):
    if not use_ui:
        return None
    sheetname = simpledialog.askstring("シート名入力", "記入用シートの名前を入力してください:", parent=root)
    return sheetname or default_name

# -------------------------
# -------------------------
# 結合ユーティリティ
# -------------------------
# 入力のベース名（拡張子なし）末尾が "…Part10" のような形式を想定
_PART_TAIL = re.compile(r"^(.*?Part)(\d+)$")

def _split_base_and_part(basename_noext: str):
    """
    入力ファイル名（拡張子なし）の末尾から "…PartN" を検出して
    (prefix_part, N) を返す。prefix_part は "…Part" の末尾まで含む。
    見つからない場合は (None, None)
    """
    m = _PART_TAIL.search(basename_noext)
    if not m:
        return None, None
    prefix_part = m.group(1)  # "…Part"
    num = int(m.group(2))
    return prefix_part, num

def _find_latest_output_for_part(outdir: str, target_basename_noext: str):
    """
    outdir 内から "*_{target_basename_noext}.xlsx" を探索し、最終更新が最新の1件を返す。
    見つからなければ None。
    例: target_basename_noext = "標的_…_Part7"
    実際のファイル名は "YYYY-mm-dd_HHMMSS_標的_…_Part7.xlsx" を想定。
    """
    pattern = os.path.join(outdir, f"*_{target_basename_noext}.xlsx")
    candidates = glob.glob(pattern)
    if not candidates:
        return None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]

_EXCEL_CELL_SAFE_CHARS = 30000

def _load_run_metadata(path: str | None) -> dict:
    if not path:
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {"metadata": data}
    except Exception as e:
        return {"metadata_load_error": f"{path}: {e}"}

def _metadata_value_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return str(value)

def _append_metadata_value(ws, key: str, value) -> None:
    text = _metadata_value_to_text(value)
    if len(text) <= _EXCEL_CELL_SAFE_CHARS:
        ws.append([key, text])
        return
    total = (len(text) + _EXCEL_CELL_SAFE_CHARS - 1) // _EXCEL_CELL_SAFE_CHARS
    for idx in range(total):
        chunk = text[idx * _EXCEL_CELL_SAFE_CHARS:(idx + 1) * _EXCEL_CELL_SAFE_CHARS]
        ws.append([f"{key} ({idx + 1}/{total})", chunk])

def _append_run_metadata_sheet(wb_out, metadata: dict) -> None:
    ws_meta = wb_out.create_sheet(title="実行条件")
    ws_meta.append(["項目", "値"])

    ordered_keys = [
        "generated_at", "merged_at", "mode", "parts", "base_excel_path", "base_name",
        "sheet_name", "input_columns", "prompt1_path", "prompt1_text",
        "prompt2_path", "prompt2_text", "apikey_path", "out_dir",
        "model_stage1", "model_stage2", "python_command",
    ]
    written = set()
    for key in ordered_keys:
        if key in metadata:
            _append_metadata_value(ws_meta, key, metadata.get(key))
            written.add(key)
    for key in sorted(k for k in metadata.keys() if k not in written):
        _append_metadata_value(ws_meta, key, metadata.get(key))

def merge_parts_into_single(outdir: str, input_base_noext: str, sheetname: str, parts: int, run_metadata_path: str | None = None) -> str:
    """
    outdir 内の "*_{…Part1}.xlsx" 〜 "*_{…PartN}.xlsx" の最新ファイルを拾い、
    1つのExcelにストリーミング結合する（ヘッダーは Part1 のみ）。
    戻り値: 出力ファイルのフルパス
    """
    prefix_part, current_num = _split_base_and_part(input_base_noext)
    if prefix_part is None:
        raise ValueError(f"入力ベース名に 'PartN' が見つかりません: {input_base_noext}")

    # 各 Part の最新出力を収集
    gathered = []
    missing = []
    for n in range(1, parts + 1):
        target_noext = f"{prefix_part}{n}"  # "…Part1" など
        p = _find_latest_output_for_part(outdir, target_noext)
        if p:
            gathered.append((n, p))
        else:
            missing.append(n)

    if missing:
        raise FileNotFoundError(
            f"以下のPart出力が見つかりませんでした: {missing}\n"
            f"フォルダ: {outdir}\n"
            f"期待パターン: '*_{prefix_part}[1..{parts}].xlsx'"
        )

    # 出力ファイル名
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    merged_name = f"{timestamp}_MERGED_{prefix_part}1-{parts}.xlsx"
    merged_path = os.path.join(outdir, merged_name)

    # write_only でストリーミング書き込み
    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title=sheetname)
    wrote_header = False

    # Part1→PartN の順に結合
    for n, src_path in sorted(gathered, key=lambda x: x[0]):
        print(f"[MERGE] Part{n}: {src_path}")
        wb_src = load_workbook(src_path, read_only=True, data_only=True)
        if sheetname not in wb_src.sheetnames:
            wb_src.close()
            raise KeyError(f"{os.path.basename(src_path)} にシート '{sheetname}' が見つかりません")
        ws_src = wb_src[sheetname]

        row_iter = ws_src.iter_rows(values_only=True)
        for idx, row in enumerate(row_iter, start=1):
            if idx == 1:
                if not wrote_header:
                    ws_out.append(list(row) if row is not None else [])
                    wrote_header = True
                # 2つ目以降のブックではヘッダー行をスキップ
            else:
                ws_out.append(list(row) if row is not None else [])
        wb_src.close()

    metadata = _load_run_metadata(run_metadata_path)
    metadata.setdefault("merged_at", datetime.now().isoformat(timespec="seconds"))
    metadata.setdefault("merge_base", input_base_noext)
    metadata.setdefault("merged_output_path", merged_path)
    _append_run_metadata_sheet(wb_out, metadata)

    wb_out.save(merged_path)
    wb_out.close()
    return merged_path

#Excel右端列検出
def rightmost_filled_col(ws, header_row=1):
    # ws.max_column はワークシートの理論上の最大列なので、
    # 右端から実際に値があるか確かめていきます。
    for col in range(ws.max_column, 0, -1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None and str(val).strip() != "":
            return col  # 列番号（1始まり）
    return 0  # 全部空なら 0


# -------------------------
# 早期リターン: 結合だけして終了（merge-only）
# -------------------------
if args.merge_only:
    if not args.outdir:
        notify_error("結合エラー", "--outdir を指定してください（例: ${workspaceFolder}/out）")
        sys.exit(1)
    if not args.merge_base:
        notify_error("結合エラー", "--merge-base を指定してください（例: 標的_…_Part10）")
        sys.exit(1)

    sheetname_for_merge = args.sheet if args.sheet else "Sheet1"
    try:
        merged_path = merge_parts_into_single(
            outdir=args.outdir,
            input_base_noext=args.merge_base,  # 末尾が PartN の拡張子なし名
            sheetname=sheetname_for_merge,
            parts=args.parts,
            run_metadata_path=args.run_metadata
        )
        notify_info("結合完了", f"Part1〜Part{args.parts} を結合しました:\n{merged_path}")
        print(f"[MERGE DONE] {merged_path}")
        sys.exit(0)
    except Exception as me:
        notify_error("結合エラー", str(me))
        sys.exit(1)

# -------------------------
# 入力・シート・プロンプトの決定（通常処理）
# -------------------------
def _fail_and_exit(msg):
    notify_error("エラー", msg)
    sys.exit(1)


def _file_sha1_short(path: str) -> str:
    try:
        h = hashlib.sha1()
        with open(path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()[:12]
    except Exception:
        return "unavailable"


def _input_columns_from_args() -> List[str]:
    cols = [c.strip() for c in (args.input_columns or []) if c and c.strip()]
    return cols if cols else list(DEFAULT_INPUT_COLUMNS)


def _has_text(value) -> bool:
    return pd.notna(value) and str(value).strip() != ""


def _row_inputs_payload(row, input_columns: List[str]) -> Tuple[List[str], str]:
    values: List[str] = []
    inputs = []
    for idx, col in enumerate(input_columns, start=1):
        raw = row[col]
        value = "" if pd.isna(raw) else str(raw)
        values.append(value)
        inputs.append({
            "placeholder": f"[入力{idx}]",
            "column": col,
            "value": value,
        })
    return values, json.dumps({"inputs": inputs}, ensure_ascii=False)


def _write_inputs_payload_file(inputs_json: str) -> str:
    tmp = tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        suffix=".json",
        prefix="copilot_inputs_",
        delete=False,
    )
    try:
        tmp.write(inputs_json)
        return tmp.name
    finally:
        tmp.close()


def _cleanup_temp_file(path: str):
    try:
        os.unlink(path)
    except Exception:
        pass


def _subprocess_env() -> dict[str, str]:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    return env

if args.input:
    wbPath = args.input
else:
    if use_ui:
        notify_info("ToDo", "特許情報が記入されたファイル (Excel) を選択してください。")
        wbPath = select_file("ファイルを選択", [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
    else:
        _fail_and_exit("--input を指定するか、UI有効で実行してください。")

if not wbPath:
    _fail_and_exit("ファイルが選択されませんでした。")
print(f"[PATH] 入力: {wbPath}")

if args.sheet:
    sheetname = args.sheet
else:
    if use_ui:
        sheetname = ask_sheetname("savedrecs")
        if not sheetname:
            notify_info("情報", "シート名が入力されませんでした。デフォルト名の「savedrecs」で実行します。")
            sheetname = "savedrecs"
    else:
        sheetname = "savedrecs"
print(f"[SHEET] {sheetname}")

# プロンプト（TXT）
if args.prompt1:
    prompt_path_stage1 = args.prompt1
else:
    if use_ui:
        notify_info("ToDo", "【1段階目】スクリーニング用のプロンプトファイル(.txt)を選択してください。")
        prompt_path_stage1 = select_file("1段階目プロンプトを選択", [("Text files", "*.txt"), ("All files", "*.*")])
        if not prompt_path_stage1:
            sys.exit("1段階目のプロンプトが選択されませんでした。終了。")
    else:
        _fail_and_exit("--prompt1 を指定するか、UI有効で実行してください。")

# 実際に使うプロンプトファイルを固定表示（誤指定切り分け用）
prompt_path_stage1 = os.path.abspath(prompt_path_stage1)
if not os.path.exists(prompt_path_stage1):
    _fail_and_exit(f"1段階目プロンプトが見つかりません: {prompt_path_stage1}")
print(f"[PROMPT1] {prompt_path_stage1} (sha1:{_file_sha1_short(prompt_path_stage1)})")

# APIキー
Gemini_APIKey = args.apikey if args.apikey else default_api_key_path
if Gemini_APIKey:
    if not os.path.exists(Gemini_APIKey):
        _fail_and_exit(f"GitHubトークンのファイルが見つかりません: {Gemini_APIKey}")
    print(f"[AUTH] token file: {os.path.abspath(Gemini_APIKey)}")
else:
    print("[AUTH] token file未指定。ログイン済みCopilotを使用します。")

# 出力先フォルダ
if args.outdir:
    outdir = args.outdir
else:
    outdir = os.path.dirname(os.path.abspath(wbPath))
os.makedirs(outdir, exist_ok=True)

# -------------------------
# 出力ファイルの用意（入力ごとに別ファイル）
# -------------------------
dir_name, base_name = os.path.split(wbPath)
file_name, file_ext = os.path.splitext(base_name)  # 例: "標的_…_Part7", ".xlsx"
formatted_datetime = datetime.now().strftime('%Y-%m-%d_%H%M%S')
wbName2 = f"{formatted_datetime}_{file_name}{file_ext}"
wbPath2 = os.path.join(outdir, wbName2)

if not os.path.exists(wbPath):
    _fail_and_exit(f"元ファイルが見つかりません: {wbPath}")

wb = load_workbook(wbPath)
wb.save(wbPath2)
print(f"[PATH] 出力: {wbPath2}")

# 右端列の取得（コピー側の中身で調べる）
ws = wb[sheetname]                 # 例: sheetname = "savedrecs"
last_col = rightmost_filled_col(ws, header_row=1)
print("右端の列番号:", last_col, "（列記号:", get_column_letter(last_col) if last_col else "-", "）")
wb.save(wbPath2)

#入力するセルの列位置
incl=last_col + 1
ws.cell(row=1, column=incl, value="1_Judgment")
ws.cell(row=1, column=incl+1, value="1_Reason")
ws.cell(row=1, column=incl+2, value="(In Case Stage1 Error: Output of LLM)")
wb.save(wbPath2)

# -------------------------
# データ読み込み
# -------------------------
try:
    df = pd.read_excel(wbPath2, sheet_name=sheetname)
except Exception as e:
    notify_error("読み込み失敗", f"{wbPath2} / {sheetname}: {e}")
    sys.exit(1)

input_columns = _input_columns_from_args()
missing_cols = [col for col in input_columns if col not in df.columns]
if missing_cols:
    _fail_and_exit(f"シート {sheetname} に必要列がありません: {', '.join(missing_cols)}")

print(f"[INPUT COLUMNS] {input_columns}")
input_column_data = df[input_columns]
non_empty_mask = input_column_data.apply(lambda col: col.map(_has_text)).any(axis=1)
iter_df = input_column_data[non_empty_mask]
print(f"[INPUT ROWS] total={len(input_column_data)} process={len(iter_df)} skipped_all_empty={len(input_column_data) - len(iter_df)}")
progress_bar = tqdm(total=len(iter_df))

notify_info("開始", "処理を開始します。")

# -------------------------
# メインループ
# -------------------------
try:
    for index, row in iter_df.iterrows():
        row_number = index + 2  # ヘッダー1行 → データは2行目〜
        row_input_values, inputs_json = _row_inputs_payload(row, input_columns)
        inputs_json_file = _write_inputs_payload_file(inputs_json)
        Abstract_text = row_input_values[0] if len(row_input_values) >= 1 else ""
        Title_text = row_input_values[1] if len(row_input_values) >= 2 else ""

        title_preview = (Title_text or Abstract_text or "")[:50]
        print(f"\n--- 処理中の行: {row_number} | 入力プレビュー: {title_preview}... ---")

        # --- 1段階目: スクリーニング ---
        print("ステージ1: ユーザの指定によりスクリーニングのみ行います...")
        stage1_cmd = [
            sys.executable, Gemini_summarize_path, "", Gemini_APIKey,
            prompt_path_stage1, "", "--inputs-json-file", inputs_json_file,
            "--model", args.model_stage1, "--stage", "screening"
        ]
        if args.web_search:
            stage1_cmd.append("--web-search")
        result_stage1 = subprocess.run(
            stage1_cmd,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            encoding='utf-8', errors='replace', env=_subprocess_env()
        )
        raw_output_stage1 = result_stage1.stdout.strip()
        raw_error_stage1 = result_stage1.stderr.strip()
        if raw_error_stage1:
            print(f"[Stage1 stderr] {raw_error_stage1}", file=sys.stderr)

        # ★ 厳格JSONバリデーション：JSON以外は即・棄却（is_relevant=False）
        is_relevant = False
        reason_text = "JSON形式エラーにより棄却"

        parsed = parse_json_input(raw_output_stage1)
        if isinstance(parsed, dict) and "_json_error" not in parsed:
        #    # JSONとして有効 → is_relevant と reason を取り出す（なければ既定）
            is_relevant = bool(parsed.get("is_relevant", False))
            reason_text = parsed.get("reason", "理由のキーが見つかりません")
            error_text= ""
        else:
            # JSONでない → 棄却。理由にエラー要約を残す
            err = parsed.get("_json_error") if isinstance(parsed, dict) else "unknown parse error"
            raw_llm_text = parsed.get("_raw_text", raw_output_stage1) if isinstance(parsed, dict) else raw_output_stage1
            if not raw_llm_text:
                raw_llm_text = raw_error_stage1
            head = raw_llm_text[:160]
            reason_text = f"非JSON出力のため棄却（Stage1）。parse_error={err}; head='{head}'"
            error_text= raw_llm_text

        print(f"[SCREENING] row={row_number} result={is_relevant}")
        print(f"[SCREENING REASON] {reason_text}")
        if error_text:
            print(f"[SCREENING RAW OUTPUT] {error_text}")

        # 1段階目の結果を書き込み
        try:
            wb = load_workbook(wbPath2)
            ws = wb[sheetname]
            ws.cell(row=row_number, column=incl, value=str(is_relevant))
            ws.cell(row=row_number, column=incl+1, value=reason_text)
            ws.cell(row=row_number, column=incl+2, value=error_text)
            wb.save(wbPath2)
            print(f"ステージ1 結果: {is_relevant} -> {incl}列目、理由を{incl+1}列目に書き込みました。")
        except Exception as e:
            print(f"[WARN] ステージ1書き込み失敗（row={row_number}）: {e}")

        # --- 2段階目: 詳細抽出 ---
        #if is_relevant:
        #    print("ステージ2: 詳細抽出処理を実行中...")
        #    result_stage2 = subprocess.run(
        #        ["python", Gemini_summarize_path, str(Abstract_text), Gemini_APIKey, prompt_path_stage2, str(Title_text), model_stage2],
        #        stdout=subprocess.PIPE, text=True, encoding='utf-8', errors='replace'
        #    )
        #    sum_text = result_stage2.stdout.strip()

         #   print(f"ステージ2 結果: {incl+3} 列目以降に書き込みます...")
            # 外部スクリプトがExcelに追記するため、こちらでは開かずに渡す
         #   subprocess.run(
         #       ["python", Split_Abstract_path, sum_text, wbPath2, str(row_number), str(incl+3), sheetname],
         #       capture_output=False, text=True, encoding='utf-8', errors='replace'
         #   )
        #else:
        #    print("ステージ2はスキップされました。")
        #    try:
        #        wb = load_workbook(wbPath2)
        #        ws = wb[sheetname]
        #        ws.cell(row=row_number, column=8, value="N/A (Skipped)")
        #        wb.save(wbPath2)
        #    except Exception as e:
        #        print(f"[WARN] ステージ2スキップ書き込み失敗（row={row_number}）: {e}")

        progress_bar.update(1)
        _cleanup_temp_file(inputs_json_file)

except Exception as e:
    print(f"[ERROR] メインループで予期せぬエラー: {e}")
finally:
    if "inputs_json_file" in locals():
        _cleanup_temp_file(inputs_json_file)
    progress_bar.close()
    notify_info("完了", "すべての処理が完了しました。")

# -------------------------
# 実行の最後で結合（merge-after 指定時のみ）
# -------------------------
if args.merge_after:
    try:
        # この実行の入力ファイル名（拡張子なし）をベースとする
        base_noext = file_name  # 例: "標的_…_Part7"
        merged_path = merge_parts_into_single(
            outdir=outdir,
            input_base_noext=base_noext,
            sheetname=sheetname,
            parts=args.parts,
            run_metadata_path=args.run_metadata
        )
        notify_info("結合完了", f"Part1〜Part{args.parts} を結合しました:\n{merged_path}")
        print(f"[MERGE DONE] {merged_path}")
    except Exception as me:
        notify_error("結合エラー", str(me))
